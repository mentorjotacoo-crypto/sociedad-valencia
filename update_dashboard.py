# -*- coding: utf-8 -*-
"""
Auto-actualizador del dashboard Sociedad Valencia (Granja Avicola Valense).

Lee los Registros de Lote (hoja CRIA-LEV) y el Consolidado Gastos Levantes
desde la carpeta de Google Drive, regenera los bloques de datos marcados con
// @@L{lote}_{TIPO}@@ dentro de index.html, y hace commit + push (GitHub Pages
redespliega automaticamente).

Uso:
    python update_dashboard.py            # actualiza, commit y push
    python update_dashboard.py --dry-run  # muestra cambios sin escribir nada
    python update_dashboard.py --no-push  # actualiza y commit, sin push

Configuracion en config_local.json (no versionado):
    source_dir, y por lote: registro, nacimiento, pollitas_registro,
    pollitas_totales, extra, color, evento.

Los cambios estructurales (lote nuevo, liquidacion) se detectan y se avisan,
pero se hacen a mano (o con Claude).
"""
import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent
INDEX = REPO / "index.html"
CONFIG = REPO / "config_local.json"
LOG = REPO / "update_log.txt"

MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
         "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
MESES_ABR = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
             "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

# Mapeo seccion de la hoja "LOTE XXXX" del Consolidado -> grupo del dashboard.
# 'PRESTACION DE SERVIC' aparece dos veces: la 1a (vacunacion) -> Vacunas,
# la 2a (galponero) -> Arrendamiento (mismo criterio del historico).
SECTION_GROUP = {
    "POLLITAS": "Pollitas",
    "ALIMENTO": "Alimento",
    "MEDICAMENTOS": "Medicamentos",
    "VACUNAS": "Vacunas",
    "CAL": "Costos comunes",
    "CISCO": "Costos comunes",
    "PAPEL": "Costos comunes",
    "GAS": "Costos comunes",
    "ARRENDAMIENTO": "Arrendamiento",
    "SEGURIDAD SOCIAL": "Costos comunes",
    "COMBUSTIBLE": "Fletes y otros",
    "PEAJES": "Fletes y otros",
    "FLETES": "Fletes y otros",
    "FLETE": "Fletes y otros",
    "SERVICIOS PUBLICOS": "Costos comunes",
    "COSTOS VARIOS": "Costos comunes",
    "GASTOS DIVERSOS": "Costos comunes",
    "VARIOS": "Costos comunes",
    "DIVERSOS": "Costos comunes",
    "ADECUACIONES": "Costos comunes",
}
GROUPS = ["Pollitas", "Alimento", "Medicamentos", "Vacunas",
          "Arrendamiento", "Servicios", "Fletes y otros", "Costos comunes"]


def log(msg):
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{stamp}] {msg}"
    print(line)
    try:
        with open(LOG, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except OSError:
        pass


def copiar_a_temp(src: Path) -> Path:
    """Copia el Excel a temp para no chocar con la sincronizacion de Drive."""
    dst = Path(tempfile.gettempdir()) / f"valense_{src.name}"
    shutil.copy2(src, dst)
    return dst


# ---------------------------------------------------------------- CRIA-LEV --
def leer_crialev(path: Path):
    """Extrae series semanales del registro de lote (hoja CRIA-LEV)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["CRIA-LEV"]
    rows = list(ws.iter_rows(min_row=1, max_row=8 + 3 * 18, max_col=26, values_only=True))
    wb.close()

    d = {"mort": [], "mort_acum": [], "kg": [], "gr_ave": [],
         "peso": [], "peso_guia_full": [], "uniformidad": None,
         "saldo": None, "semanas": 0, "consumo_total": 0}
    for w in range(18):
        mrow = rows[7 + 3 * w] if len(rows) > 7 + 3 * w else None
        krow = rows[9 + 3 * w] if len(rows) > 9 + 3 * w else None
        if mrow is None:
            break
        # guia de peso disponible para todas las semanas (col 24)
        guia = mrow[23]
        if isinstance(guia, (int, float)) and guia > 0:
            d["peso_guia_full"].append(int(round(guia)))
        kg = mrow[15] if isinstance(mrow[15], (int, float)) else 0
        if not kg or kg <= 0:
            continue  # semana sin datos
        d["semanas"] = w + 1
        d["mort"].append(int(mrow[9] or 0))
        d["mort_acum"].append(round(float(mrow[14] or 0), 2))
        d["kg"].append(int(kg))
        d["consumo_total"] += int(kg)
        gr = mrow[18]
        d["gr_ave"].append(round(float(gr), 1) if isinstance(gr, (int, float)) and 0 < gr < 500 else None)
        peso = mrow[22]
        d["peso"].append(round(float(peso), 1) if isinstance(peso, (int, float)) and peso > 0 else None)
        if isinstance(mrow[12], (int, float)) and mrow[12] > 0:
            d["saldo"] = int(mrow[12])
        if krow and isinstance(krow[22], (int, float)) and krow[22] > 0:
            d["uniformidad"] = round(float(krow[22]), 1)

    # metricas finales
    d["mort_total"] = sum(d["mort"])
    d["mort_pct"] = d["mort_acum"][-1] if d["mort_acum"] else 0
    pesos_validos = [(i, p) for i, p in enumerate(d["peso"]) if p]
    if pesos_validos:
        i, p = pesos_validos[-1]
        d["peso_final"] = int(round(p))
        guia = d["peso_guia_full"][i] if i < len(d["peso_guia_full"]) else None
        d["peso_guia_sem"] = guia
        d["cumpl"] = round((p - guia) / guia * 100, 1) if guia else None
    else:
        d["peso_final"] = d["peso_guia_sem"] = d["cumpl"] = None
    return d


# ------------------------------------------------------------------ GASTOS --
def leer_gastos(consolidado: Path, lote: str, nacimiento: date, pollitas: int):
    """Extrae gastos semanales por grupo de la hoja LOTE {lote} del Consolidado."""
    wb = openpyxl.load_workbook(consolidado, data_only=True, read_only=True)
    ws = wb[f"LOTE {lote}"]
    section, presta = None, 0
    weekly = {g: [0] * 18 for g in GROUPS}
    totals = {g: 0 for g in GROUPS}
    for row in ws.iter_rows(min_row=1, max_row=1000, max_col=6, values_only=True):
        a, fecha, _fact, _nombre, valor, _cant = (list(row) + [None] * 6)[:6]
        if isinstance(a, str) and a.strip():
            key = a.strip().upper()
            if key.startswith("INGRESOS") or key.startswith("MORTALIDAD"):
                section = None
                continue
            if key.startswith("PRESTACION DE SERVIC"):
                presta += 1
                section = "Vacunas" if presta == 1 else "Arrendamiento"
                continue
            matched = next((g for sk, g in SECTION_GROUP.items() if key.startswith(sk)), None)
            if matched:
                section = matched
            elif not key.startswith("LEVANTE") and not key[0].isdigit():
                section = None
            continue
        if section is None or valor is None or not isinstance(valor, (int, float)) or valor == 0:
            continue
        fdate = None
        if isinstance(fecha, datetime):
            fdate = fecha.date()
        elif isinstance(fecha, str) and fecha.strip() and fecha.strip().upper() != "TOTAL":
            parts = [p for p in fecha.replace("/", "-").split("-") if p.strip().isdigit()]
            if len(parts) >= 3:
                dd, mm, yy = int(parts[0]), int(parts[1]), int(parts[2])
                if yy < 100:
                    yy += 2000
                try:
                    fdate = date(yy, mm, dd)
                except ValueError:
                    pass
        if fdate is None:
            continue
        wk = max(0, min(17, (fdate - nacimiento).days // 7))
        weekly[section][wk] += int(round(valor))
        totals[section] += int(round(valor))
    wb.close()
    return {"pollitas": pollitas, "group_totals": totals,
            "weekly_by_group": weekly, "grand_total": sum(totals.values())}


def lotes_liquidados(consolidado: Path):
    """Lee la tabla de liquidaciones de la hoja Consolidado."""
    wb = openpyxl.load_workbook(consolidado, data_only=True, read_only=True)
    ws = wb["Consolidado"]
    liq = set()
    for row in ws.iter_rows(min_row=20, max_row=60, max_col=7, values_only=True):
        desc, estado = row[1], row[5]
        if isinstance(desc, str) and isinstance(estado, str) and "liquidado" in estado.lower():
            m = re.search(r"(\d{4})", desc)
            if m:
                liq.add(m.group(1))
    wb.close()
    return liq


# --------------------------------------------------------------- RENDERING --
def js_int_arr(a):
    return "[" + ",".join(str(int(v)) for v in a) + "]"


def js_f1_arr(a):
    return "[" + ",".join("null" if v is None else f"{v:.1f}" for v in a) + "]"


def js_f2_arr(a):
    return "[" + ",".join("null" if v is None else f"{v:g}" if v == int(v) else f"{v:.2f}" for v in a) + "]"


def generar_lineas(lote: str, cfg: dict, cl: dict, ga: dict):
    """Genera las lineas de datos que reemplazan a las marcadas con @@."""
    gt = ga["group_totals"]
    extra = f"extra:{cfg.get('extra', 0)}, " if cfg.get("extra", 0) else ""
    consumo = (
        f"  '{lote}':{{label:'L.{lote}',color:'{cfg['color']}',pollitas:{cfg['pollitas_registro']},"
        f"kg_sem:{js_int_arr(cl['kg'])},gr_ave:{js_f1_arr(cl['gr_ave'])},"
        f"peso_real:{js_f1_arr(cl['peso'])},peso_guia:{js_int_arr(cl['peso_guia_full'])},"
        f"mort:{js_int_arr(cl['mort'])},consumo_total:{cl['consumo_total']},"
        f"utilidad:null,mortalidad_pct:{cl['mort_pct']:.2f},activo:true}},"
    )
    mpstats = (
        f"  '{lote}': {{pollitas:{cfg['pollitas_registro']}, {extra}mort_total:{cl['mort_total']}, "
        f"mort_pct:{cl['mort_pct']:.2f}, peso_final:{cl['peso_final']}, peso_guia:{cl['peso_guia_sem']}, "
        f"cumpl:{cl['cumpl']}, uniformidad:{cl['uniformidad']:.1f}, semanas:{cl['semanas']}, "
        f"evento:'{cfg.get('evento', 'Sin eventos criticos')}', activo:true}},"
    )
    graw_dict = {"pollitas": cfg["pollitas_totales"], "group_totals": gt,
                 "weekly_by_group": ga["weekly_by_group"], "grand_total": ga["grand_total"]}
    return {
        "MORT_W": f"    '{lote}': {js_int_arr(cl['mort'])},",
        "MORT_A": f"    '{lote}': {js_f2_arr(cl['mort_acum'])},",
        "PESO_R": f"    '{lote}': {js_f1_arr(cl['peso'])},",
        "PESO_G": f"    '{lote}': {js_int_arr(cl['peso_guia_full'])},",
        "COSTS_CAT": (
            f"    '{lote}': {{Pollitas:{gt['Pollitas']},Alimento:{gt['Alimento']},'M.Obra':0,"
            f"Arriendo:{gt['Arrendamiento']},Vacunas:{gt['Vacunas']},Medicam:{gt['Medicamentos']},"
            f"Fletes:{gt['Fletes y otros']},Otros:{gt['Costos comunes']}}},"
        ),
        "TOTALS": (
            f"    '{lote}': {{costo:{ga['grand_total']}, ingreso:0, utilidad:0, aves:0, "
            f"pollitas:{cfg['pollitas_totales']}}},"
        ),
        "MPSTATS": mpstats,
        "CONSUMO": consumo,
        "GRAW": f'"{lote}":{json.dumps(graw_dict, ensure_ascii=False, separators=(",", ":"))},',
    }


def aplicar(html: str, lote: str, lineas: dict):
    """Reemplaza cada linea marcada con // @@L{lote}_{TIPO}@@ conservando el marcador."""
    faltantes = []
    for kind, nueva in lineas.items():
        marker = f"@@L{lote}_{kind}@@"
        pat = re.compile(rf"^.*// {re.escape(marker)}.*$", re.M)
        if not pat.search(html):
            faltantes.append(marker)
            continue
        html = pat.sub(lambda _m: f"{nueva} // {marker}", html)
    return html, faltantes


def actualizar_corte(html: str, corte: date):
    badge = f"Corte {MESES_ABR[corte.month - 1]} {corte.day}, {corte.year}"
    heading = f"Corte {MESES[corte.month - 1]} {corte.day}, {corte.year}"
    html = re.sub(r'(id="corte-badge">)[^<]*(</span>)', lambda m: m.group(1) + badge + m.group(2), html)
    html = re.sub(r'(id="corte-heading">)[^<]*(</span>)', lambda m: m.group(1) + heading + m.group(2), html)
    return html


# --------------------------------------------------------------------- GIT --
def git(*args):
    r = subprocess.run(["git", "-C", str(REPO), *args],
                       capture_output=True, text=True, encoding="utf-8")
    if r.returncode != 0:
        raise RuntimeError(f"git {' '.join(args)} fallo: {r.stderr.strip()}")
    return r.stdout.strip()


# -------------------------------------------------------------------- MAIN --
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="no escribe ni hace commit")
    ap.add_argument("--no-push", action="store_true", help="commit sin push")
    args = ap.parse_args()

    if not CONFIG.exists():
        log("ERROR: falta config_local.json")
        return 1
    cfg = json.loads(CONFIG.read_text(encoding="utf-8"))
    src_dir = Path(cfg["source_dir"])
    if not src_dir.exists():
        log(f"ERROR: carpeta de Drive no disponible: {src_dir}")
        return 1

    consolidado = copiar_a_temp(src_dir / "Consolidado Gastos Levantes.xlsx")
    liquidados = lotes_liquidados(consolidado)

    html = INDEX.read_text(encoding="utf-8")
    corte_max = None

    for lote, lc in cfg["lotes_activos"].items():
        if lote in liquidados:
            log(f"AVISO: el lote {lote} figura LIQUIDADO en el Consolidado. "
                f"Actualiza el dashboard manualmente (estructura) y retiralo de config_local.json.")
            continue
        registro = copiar_a_temp(src_dir / lc["registro"])
        nacimiento = date.fromisoformat(lc["nacimiento"])
        cl = leer_crialev(registro)
        if cl["semanas"] == 0:
            log(f"AVISO: lote {lote} sin semanas con datos; se omite.")
            continue
        ga = leer_gastos(consolidado, lote, nacimiento, lc["pollitas_totales"])
        lineas = generar_lineas(lote, lc, cl, ga)
        html, faltantes = aplicar(html, lote, lineas)
        for m in faltantes:
            log(f"AVISO: marcador {m} no encontrado en index.html")
        corte_lote = nacimiento + timedelta(days=cl["semanas"] * 7 - 1)
        corte_max = max(corte_max, corte_lote) if corte_max else corte_lote
        log(f"Lote {lote}: sem {cl['semanas']}, mort {cl['mort_pct']:.2f}%, "
            f"peso {cl['peso_final']}g, gastos ${ga['grand_total']:,}")

    if corte_max:
        html = actualizar_corte(html, corte_max)

    original = INDEX.read_text(encoding="utf-8")
    if html == original:
        log("Sin cambios: el dashboard ya esta al dia.")
        return 0

    if args.dry_run:
        import difflib
        diff = difflib.unified_diff(original.splitlines(), html.splitlines(),
                                    "actual", "nuevo", lineterm="", n=0)
        print("\n".join(list(diff)[:120]))
        log("DRY-RUN: habia cambios (no se escribio nada).")
        return 0

    INDEX.write_text(html, encoding="utf-8")
    log("index.html actualizado.")

    if git("status", "--porcelain", "index.html"):
        git("add", "index.html")
        git("commit", "-m", f"Auto-update {datetime.now():%Y-%m-%d %H:%M}")
        log("Commit creado.")
        if not args.no_push:
            git("push", "origin", "main")
            log("Push a GitHub OK: Pages redesplegara en ~1 min.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
